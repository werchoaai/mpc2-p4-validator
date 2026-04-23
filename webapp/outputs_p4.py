"""
P4-specific Excel output: Overview / Details / Raw workbooks in legacy format.

Ported from Codex's Project-4 Streamlit app. Reuses the A1 parser's
`Measurement` core via `process_measurement(...)`.
"""
from __future__ import annotations

import io
from datetime import datetime, time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OVERVIEW_HEADERS = [
    "Messung", "Datum", "Uhrzeit", "Messzelle", "Versuchsnr.", "order no.",
    "material", "condition", "designation", "LiMi ", "Messflaeche [mm2]",
    "EPR-Loesung", "", "H2SO4 [ml/l]", "HCl [ml/l]", "KSCN [mM]",
    "Messgeschw. [mV/s]", "Vertexpotential [mV]",
    "OCP [mV]", "Ja [mA/cm2]", "Jr [mA/cm2]", "Jr/Ja",
    "Qa [As]", "Qr [As]", "Qr/Qa",
    "comments", "operator",
]

DETAIL_HEADERS = [
    "Zeit [s]", "Potenzial [V]", "I [A]", "J [A/m2]",
    "Potenzial [mV]", "J [mA/cm2]", "delta t [s]", "delta Q [As]",
]

MATERIAL_LABEL_MAP = {"1.4466": "25-22-2"}

COMMENT_BY_DESIGNATION = {
    "L1": "left / top", "L2": "left / center", "L3": "left / bottom",
    "R1": "right / top", "R2": "right / center", "R3": "right / bottom",
    "P1": "reference", "REFERENZ P1": "reference",
}


def _safe_sheet_name(name: str) -> str:
    safe = "".join(c for c in name if c not in r"[]:*?/\\")
    return safe[:31] or "Sheet"


def _default_material_label(raw_material):
    if not raw_material:
        return "material"
    return MATERIAL_LABEL_MAP.get(raw_material, raw_material)


def _detail_sheet_name(measurement, material_label: str) -> str:
    messung = measurement.filename_meta.messung_id or "0000"
    display_id = messung.lstrip("0") or "0"
    designation = measurement.filename_meta.probenbez or "sample"
    return _safe_sheet_name(f"{display_id}_{material_label}_{designation}")


def _summary_row(measurement, defaults: dict, index: int) -> dict:
    """Build one overview row from a parsed Measurement + defaults."""
    fm = measurement.filename_meta
    analysis = measurement.analysis
    material_label = _default_material_label(fm.material)
    designation = (fm.probenbez or "").strip()
    comment = COMMENT_BY_DESIGNATION.get(designation.upper(), defaults.get("comment", ""))

    date_str = defaults.get("date") or datetime.now().strftime("%d.%m.%Y")
    time_str = defaults.get("time") or datetime.now().strftime("%H:%M")

    # Real attribute names on DLEPRResult:
    #   ja_ma_cm2, jr_ma_cm2, qa_as, qr_as, jr_ja, qr_qa, ruhepotential_mv
    return {
        "index": index,
        "Messung": fm.messung_id or "",
        "Datum": date_str,
        "Uhrzeit": time_str,
        "Messzelle": defaults.get("cell", "FC"),
        "Versuchsnr.": defaults.get("experiment_no", ""),
        "order no.": defaults.get("order_no", ""),
        "material": material_label,
        "condition": getattr(fm, "zustand", None) or "",
        "designation": designation,
        "LiMi ": defaults.get("limit", ""),
        "Messflaeche [mm2]": (measurement.had_meta.probenflaeche_mm2
                              if getattr(measurement, "had_meta", None) else ""),
        "EPR-Loesung": defaults.get("solution_label", ""),
        "": "",
        "H2SO4 [ml/l]": defaults.get("h2so4", 146),
        "HCl [ml/l]": defaults.get("hcl", 120),
        "KSCN [mM]": defaults.get("kscn", 1),
        "Messgeschw. [mV/s]": defaults.get("scan_rate", 1.67),
        "Vertexpotential [mV]": defaults.get("vertex_mv", 390),
        "OCP [mV]": round(analysis.ruhepotential_mv, 1) if analysis.ruhepotential_mv is not None else "",
        "Ja [mA/cm2]": round(analysis.ja_ma_cm2, 3) if analysis.ja_ma_cm2 is not None else "",
        "Jr [mA/cm2]": round(analysis.jr_ma_cm2, 3) if analysis.jr_ma_cm2 is not None else "",
        "Jr/Ja": round(analysis.jr_ja, 4) if analysis.jr_ja is not None else "",
        "Qa [As]": round(analysis.qa_as, 4) if analysis.qa_as is not None else "",
        "Qr [As]": round(analysis.qr_as, 4) if analysis.qr_as is not None else "",
        "Qr/Qa": round(analysis.qr_qa, 4) if analysis.qr_qa is not None else "",
        "comments": comment,
        "operator": defaults.get("operator", ""),
        "split_method": analysis.split_method,
        "split_index": analysis.split_index,
    }


def _header_style(cell):
    cell.font = Font(bold=True, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor="FF4C4C4C")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_overview_workbook(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    for col, header in enumerate(OVERVIEW_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _header_style(cell)
    for r_idx, row in enumerate(rows, start=2):
        for col, header in enumerate(OVERVIEW_HEADERS, start=1):
            ws.cell(row=r_idx, column=col, value=row.get(header, ""))
    for col in range(1, len(OVERVIEW_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_detail_workbook(measurements, rows) -> bytes:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    for col, header in enumerate(OVERVIEW_HEADERS, start=1):
        cell = summary_ws.cell(row=1, column=col, value=header)
        _header_style(cell)
    for r_idx, row in enumerate(rows, start=2):
        for col, header in enumerate(OVERVIEW_HEADERS, start=1):
            summary_ws.cell(row=r_idx, column=col, value=row.get(header, ""))

    for m, row in zip(measurements, rows):
        material_label = _default_material_label(m.filename_meta.material)
        sheet_name = _detail_sheet_name(m, material_label)
        ws = wb.create_sheet(sheet_name)
        for col, header in enumerate(DETAIL_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            _header_style(cell)
        asc = m.asc
        probenflaeche_mm2 = (m.had_meta.probenflaeche_mm2
                              if getattr(m, "had_meta", None) else 1.0)
        area_m2 = (probenflaeche_mm2 or 1.0) / 1e6
        area_cm2 = (probenflaeche_mm2 or 100.0) / 100
        for r_idx, (t, v, i) in enumerate(
            zip(asc.time_s, asc.potential_V, asc.current_A), start=2
        ):
            ws.cell(row=r_idx, column=1, value=round(t, 3))
            ws.cell(row=r_idx, column=2, value=round(v, 6))
            ws.cell(row=r_idx, column=3, value=round(i, 9))
            ws.cell(row=r_idx, column=4, value=round(i / area_m2, 4))
            ws.cell(row=r_idx, column=5, value=round(v * 1000, 3))
            ws.cell(row=r_idx, column=6, value=round((i / area_cm2) * 1000, 4))
        for col in range(1, len(DETAIL_HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_raw_workbook(measurements) -> bytes:
    wb = Workbook()
    vorlage = wb.active
    vorlage.title = "Vorlage"
    for col, header in enumerate(DETAIL_HEADERS, start=1):
        cell = vorlage.cell(row=1, column=col, value=header)
        _header_style(cell)
    vorlage.cell(row=2, column=1,
                 value="(Vorlage - Formatreferenz)").font = Font(italic=True, color="FF808080")

    for m in measurements:
        material_label = _default_material_label(m.filename_meta.material)
        sheet_name = _detail_sheet_name(m, material_label)
        ws = wb.create_sheet(sheet_name)
        for col, header in enumerate(DETAIL_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            _header_style(cell)
        asc = m.asc
        probenflaeche_mm2 = (m.had_meta.probenflaeche_mm2
                              if getattr(m, "had_meta", None) else 1.0)
        area_m2 = (probenflaeche_mm2 or 1.0) / 1e6
        area_cm2 = (probenflaeche_mm2 or 100.0) / 100
        for r_idx, (t, v, i) in enumerate(
            zip(asc.time_s, asc.potential_V, asc.current_A), start=2
        ):
            ws.cell(row=r_idx, column=1, value=round(t, 3))
            ws.cell(row=r_idx, column=2, value=round(v, 6))
            ws.cell(row=r_idx, column=3, value=round(i, 9))
            ws.cell(row=r_idx, column=4, value=round(i / area_m2, 4))
            ws.cell(row=r_idx, column=5, value=round(v * 1000, 3))
            ws.cell(row=r_idx, column=6, value=round((i / area_cm2) * 1000, 4))
        for col in range(1, len(DETAIL_HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
